import openpyxl
import os
import sys

# Set encoding for output
sys.stdout.reconfigure(encoding='utf-8')

CONFIG_PATH = 'Cau_hinh.xlsx'
if not os.path.exists(CONFIG_PATH):
    print(f"Error: {CONFIG_PATH} not found")
else:
    try:
        wb = openpyxl.load_workbook(CONFIG_PATH, data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        if 'Cấu hình' in wb.sheetnames:
            ws = wb['Cấu hình']
            print("Cấu hình sheet found")
            # Print some values to verify
            print(f"B1 (w_r): {ws.cell(row=1, column=2).value}")
            print(f"B10 (today_excel): {ws.cell(row=10, column=2).value}")
            
            # Check for NaN or None in critical spots
            for i in range(1, 7):
                print(f"Row {i}, Col 2: {ws.cell(row=i, column=2).value}")
        else:
            print("Error: 'Cấu hình' sheet not found")
    except Exception as e:
        print(f"Error reading workbook: {e}")
