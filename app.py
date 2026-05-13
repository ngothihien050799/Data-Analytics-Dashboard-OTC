from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import io
import math

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

app = Flask(__name__, static_folder='dist', static_url_path='/')
app.config['JSON_SORT_KEYS'] = False
CORS(app)

CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Cau_hinh.xlsx')

def style_excel(writer, df_tonghop, df_chamdiem):
    workbook = writer.book
    groups_tonghop = [
        ("THÔNG TIN KH", 2, "EBF5FF", "2563EB"),
        ("LỊCH SỬ MUA HÀNG", 6, "ECFDF5", "059669"),
        ("HOẠT ĐỘNG GHÉ THĂM", 2, "F5F3FF", "7C3AED"),
        ("PHÂN LOẠI & DOANH SỐ THÁNG", 3, "FFFBEB", "D97706"),
        ("CHỈ SỐ THIẾU HỤT", 3, "FEF2F2", "DC2626"),
        ("XU HƯỚNG & HÀNH ĐỘNG", 4, "ECFEFF", "0891B2")
    ]
    groups_chamdiem = [
        ("THÔNG TIN KH", 3, "EBF5FF", "2563EB"),
        ("ĐIỂM THÀNH PHẦN RFM", 6, "FFF1F2", "E11D48"),
        ("TỔNG HỢP", 1, "FFFBEB", "D97706"),
        ("HOẠT ĐỘNG", 3, "F5F3FF", "7C3AED"),
        ("CHỌN", 1, "F3F4F6", "4B5563")
    ]
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    def apply_styles(sheet_name, groups, df):
        if sheet_name not in workbook.sheetnames: return
        ws = workbook[sheet_name]
        ws.insert_rows(1)
        col_idx = 1
        for group_name, span, bg_color, text_color in groups:
            start_col = col_idx
            end_col = col_idx + span - 1
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col)
            cell.value = group_name
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True, color=text_color)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            for c in range(start_col, end_col + 1):
                sub_cell = ws.cell(row=2, column=c)
                sub_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                sub_cell.font = Font(bold=True, color=text_color)
                sub_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                sub_cell.border = thin_border
                ws.cell(row=1, column=c).border = thin_border
            col_idx += span
        ws.freeze_panes = "A3"
        for col in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        for r in range(3, len(df) + 3):
            for c in range(1, len(df.columns) + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                col_name = df.columns[c - 1]
                cell.border = thin_border
                if val is None: continue
                if isinstance(val, (int, float)):
                    if "Điểm" in col_name or col_name == "ĐIỂM TỔNG": cell.number_format = '0.00'
                    elif "DS" in col_name or "doanh số" in col_name.lower() or col_name == "Chu kỳ TB": cell.number_format = '#,##0'
                if col_name == 'Hạng KH':
                    if val == 'VIP': cell.fill, cell.font = PatternFill(start_color='FEF08A', end_color='FEF08A', fill_type='solid'), Font(color='854D0E', bold=True)
                    elif val == 'GOLD': cell.fill, cell.font = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid'), Font(color='B45309', bold=True)
                    elif val == 'NORMAL': cell.fill, cell.font = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid'), Font(color='334155', bold=True)
                    elif val == 'NEW': cell.fill, cell.font = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid'), Font(color='1E40AF', bold=True)
                elif col_name == 'Trạng thái hoạt động':
                    if val == 'Hoạt động': cell.fill, cell.font = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid'), Font(color='166534')
                    elif val == 'Cảnh báo': cell.fill, cell.font = PatternFill(start_color='FEF08A', end_color='FEF08A', fill_type='solid'), Font(color='854D0E')
                    elif val == 'Ngủ đông': cell.fill, cell.font = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'), Font(color='991B1B')
                    elif val == 'Chưa mua': cell.fill, cell.font = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid'), Font(color='374151')
                elif col_name in ['Call thiếu', 'DS thiếu', 'Call còn lại']:
                    try:
                        num = float(val)
                        if num > 0: cell.fill, cell.font = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'), Font(color='991B1B')
                        else: cell.fill, cell.font = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'), Font(color='065F46')
                    except: pass
                elif col_name == 'Số ngày chưa gặp':
                    try:
                        num = float(val)
                        if num > 14: cell.fill, cell.font = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'), Font(color='991B1B')
                        else: cell.fill, cell.font = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'), Font(color='065F46')
                    except: pass
                elif ("Điểm" in col_name or col_name == "ĐIỂM TỔNG") and isinstance(val, (int, float)):
                    s = max(0, min(1, float(val)))
                    if s >= 0.75: cell.fill, cell.font = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid'), Font(color='FFFFFF', bold=True)
                    elif s >= 0.5: cell.fill, cell.font = PatternFill(start_color='F97316', end_color='F97316', fill_type='solid'), Font(color='FFFFFF', bold=True)
                    elif s >= 0.25: cell.fill, cell.font = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid'), Font(color='FFFFFF', bold=True)
                    else: cell.fill, cell.font = PatternFill(start_color='64748B', end_color='64748B', fill_type='solid'), Font(color='FFFFFF', bold=True)
    apply_styles('7_TongHop', groups_tonghop, df_tonghop)
    apply_styles('Cham_diem_KH', groups_chamdiem, df_chamdiem)

def excel_date_to_datetime(excel_date):
    if pd.isna(excel_date):
        return None
    return datetime(1899, 12, 30) + timedelta(days=int(excel_date))

def process_data(input_file_path):
    # Load Configuration using openpyxl to avoid pandas header offset issues
    wb = openpyxl.load_workbook(CONFIG_PATH, data_only=True)
    ws = wb['Cấu hình']
    
    def parse_pct(val):
        if isinstance(val, str) and '%' in val:
            return float(val.replace('%', '')) / 100
        return float(val) if val is not None else 0.0

    w_r = parse_pct(ws.cell(row=1, column=2).value)
    w_f = parse_pct(ws.cell(row=2, column=2).value)
    w_m = parse_pct(ws.cell(row=3, column=2).value)
    w_rank = parse_pct(ws.cell(row=4, column=2).value)
    w_call = parse_pct(ws.cell(row=5, column=2).value)
    w_cycle = parse_pct(ws.cell(row=6, column=2).value)
    
    # System Params
    today_excel = ws.cell(row=10, column=2).value
    # Use actual today's date instead of config date to prevent confusion
    today = datetime.now()
    
    max_recency = float(ws.cell(row=12, column=2).value or 120)
    f_window_months = float(ws.cell(row=13, column=2).value or 4)
    max_frequency = float(ws.cell(row=14, column=2).value or 12)
    max_monetary = float(ws.cell(row=15, column=2).value or 18000000)
    max_call_days = float(ws.cell(row=16, column=2).value or 120)
    
    # Rank thresholds
    threshold_vip = float(ws.cell(row=20, column=2).value or 18000000)
    threshold_gold = float(ws.cell(row=21, column=2).value or 6000000)
    threshold_normal = float(ws.cell(row=22, column=2).value or 1000000)
    ds_window_months = float(ws.cell(row=23, column=2).value or 6)
    
    # Load Input Data
    with pd.ExcelFile(input_file_path, engine='openpyxl') as xl:
        df_nhanvien = pd.read_excel(xl, '1_NhanVien', dtype=str)
        df_kh = pd.read_excel(xl, '2_KHTrongTam')
        df_orders = pd.read_excel(xl, '3_DonHang')
        df_calls = pd.read_excel(xl, '4_Call')
        df_freq = pd.read_excel(xl, '5_FrequencyF')
    
    # Ensure date formats
    df_orders['Ngày đặt'] = pd.to_datetime(df_orders['Ngày đặt'])
    df_calls['Thời gian checkin'] = pd.to_datetime(df_calls['Thời gian checkin'])
    
    # Calculate 6-month window
    start_date_6m = today - timedelta(days=30*ds_window_months)
    
    # --- Processing 7_TongHop ---
    
    # Order statistics
    if 'Số lượng' not in df_orders.columns or 'Đơn giá' not in df_orders.columns:
        df_orders['Revenue'] = 0
    else:
        df_orders['Số lượng'] = pd.to_numeric(df_orders['Số lượng'], errors='coerce').fillna(0)
        df_orders['Đơn giá'] = pd.to_numeric(df_orders['Đơn giá'], errors='coerce').fillna(0)
        df_orders['Revenue'] = df_orders['Số lượng'] * df_orders['Đơn giá']
        
    df_orders_6m = df_orders[df_orders['Ngày đặt'] >= start_date_6m]
    
    current_month = today.month
    current_year = today.year
    prev_month = 12 if current_month == 1 else current_month - 1
    prev_year = current_year - 1 if current_month == 1 else current_year

    order_stats = []
    for makh, group in df_orders_6m.groupby('Mã KH'):
        sorted_dates = sorted(group['Ngày đặt'].dt.date.unique(), reverse=True)
        last_buy = sorted_dates[0] if sorted_dates else None
        
        diff1 = (sorted_dates[0] - sorted_dates[1]).days if len(sorted_dates) > 1 else None
        diff2 = (sorted_dates[1] - sorted_dates[2]).days if len(sorted_dates) > 2 else None
        
        ds_thang = group[(group['Ngày đặt'].dt.month == current_month) & (group['Ngày đặt'].dt.year == current_year)]['Revenue'].sum()
        ds_thang_truoc = group[(group['Ngày đặt'].dt.month == prev_month) & (group['Ngày đặt'].dt.year == prev_year)]['Revenue'].sum()
        
        if ds_thang > ds_thang_truoc: xu_huong = "📈 Tăng"
        elif ds_thang < ds_thang_truoc: xu_huong = "📉 Giảm"
        else: xu_huong = "➡ Ổn định"
        
        days_since = (today.date() - last_buy).days if last_buy else None
        
        chu_ky_tb = None
        if diff1 is not None and diff2 is not None: 
            chu_ky_tb = math.ceil((diff1 + diff2) / 2)
        
        du_bao = None
        if chu_ky_tb is not None:
            du_bao = (today.date() + timedelta(days=chu_ky_tb)).strftime('%d/%m/%Y')
            
        if days_since is None: trang_thai = "Chưa mua"
        elif days_since <= 60: trang_thai = "Hoạt động"
        elif days_since <= 120: trang_thai = "Cảnh báo"
        else: trang_thai = "Ngủ đông"
        
        order_stats.append({
            'Mã KH': makh,
            'Ngày mua cuối': last_buy,
            'Số lần mua trong 6 tháng': len(sorted_dates),
            'Tổng doanh số 6 tháng': group['Revenue'].sum(),
            'Tổng DS tháng': ds_thang,
            'Tổng DS tháng trước': ds_thang_truoc,
            'Xu hướng mua': xu_huong,
            'Số ngày mua gần nhất': days_since,
            'Số ngày mua gần thứ 2': diff1,
            'Số ngày mua gần thứ 3': diff2,
            'Chu kỳ TB': chu_ky_tb,
            'Dự báo ngày mua tiếp': du_bao,
            'Trạng thái hoạt động': trang_thai
        })
    df_order_summary = pd.DataFrame(order_stats)
    
    # Call statistics
    call_stats = []
    for makh, group in df_calls.groupby('Mã KH'):
        last_call = group['Thời gian checkin'].max()
        # Handle cases where all calls have no timestamp (NaT)
        days_since_call = (today - last_call).days if pd.notna(last_call) else None
        
        call_stats.append({
            'Mã KH': makh,
            'Số lần đã gặp': len(group),
            'Số ngày chưa gặp': days_since_call,
            'Ngày gặp cuối': last_call.strftime('%d/%m/%Y') if pd.notna(last_call) else None
        })
    df_call_summary = pd.DataFrame(call_stats)
    
    # Merge everything into 7_TongHop
    df_tonghop = df_kh[['Mã KH', 'Tên KH']].copy()
    df_tonghop = df_tonghop.merge(df_order_summary, on='Mã KH', how='left')
    df_tonghop = df_tonghop.merge(df_call_summary, on='Mã KH', how='left')
    
    # Ranking
    def get_rank(ds):
        if ds >= threshold_vip: return 'VIP'
        if ds >= threshold_gold: return 'GOLD'
        if ds >= threshold_normal: return 'NORMAL'
        return 'NEW'
    
    df_tonghop['Hạng KH'] = df_tonghop['Tổng doanh số 6 tháng'].fillna(0).apply(get_rank)
    
    # FrequencyF merge (Group by Mã KHTC)
    if 'F' in df_freq.columns:
        df_freq['Call'] = df_freq[['Call', 'F']].max(axis=1)
        
    if 'Mã KHTC' in df_freq.columns:
        df_freq_grouped = df_freq.groupby('Mã KHTC').agg({
            'Doanh số KH': 'max',
            'Call': 'sum'
        }).reset_index().rename(columns={'Mã KHTC': 'Mã KH', 'Doanh số KH': 'DS mục tiêu', 'Call': 'Call mục tiêu'})
    else:
        df_freq_grouped = pd.DataFrame(columns=['Mã KH', 'DS mục tiêu', 'Call mục tiêu'])
        
    df_tonghop = df_tonghop.merge(df_freq_grouped, on='Mã KH', how='left')
    
    # Fill Nans for missing columns after merge
    df_tonghop['Tổng DS tháng'] = df_tonghop['Tổng DS tháng'].fillna(0)
    df_tonghop['Xu hướng mua'] = df_tonghop['Xu hướng mua'].fillna("Chưa mua")
    df_tonghop['Trạng thái hoạt động'] = df_tonghop['Trạng thái hoạt động'].fillna("Chưa mua")
    df_tonghop['Số lần mua trong 6 tháng'] = df_tonghop['Số lần mua trong 6 tháng'].fillna(0)
    df_tonghop['Tổng doanh số 6 tháng'] = df_tonghop['Tổng doanh số 6 tháng'].fillna(0)
    
    df_tonghop['DS thiếu'] = df_tonghop['DS mục tiêu'].fillna(0) - df_tonghop['Tổng DS tháng']
    df_tonghop['Call thiếu'] = (df_tonghop['Call mục tiêu'].fillna(0) - df_tonghop['Số lần đã gặp'].fillna(0)).apply(lambda x: max(0, x))
    
    # Clean column names to match sample
    df_tonghop = df_tonghop.rename(columns={'Tên KH': 'Tên khách hàng'})
    
    # --- Processing Cham_diem_KH ---
    df_scoring = df_tonghop.copy()
    
    # Scoring Logic
    df_scoring['Điểm R'] = df_scoring['Số ngày mua gần nhất'].apply(lambda x: max(0, 1 - x/max_recency) if pd.notna(x) else 0)
    df_scoring['Điểm F'] = df_scoring['Số lần mua trong 6 tháng'].apply(lambda x: min(1, x/max_frequency) if pd.notna(x) else 0)
    df_scoring['Điểm M'] = df_scoring['Tổng doanh số 6 tháng'].apply(lambda x: min(1, x/max_monetary) if pd.notna(x) else 0)
    
    rank_points = {'VIP': 1.0, 'GOLD': 0.8, 'NORMAL': 0.6, 'NEW': 0.4}
    df_scoring['Điểm Hạng'] = df_scoring['Hạng KH'].map(rank_points)
    
    df_scoring['Điểm Call'] = df_scoring['Số ngày chưa gặp'].apply(lambda x: max(0, 1 - x/max_call_days) if pd.notna(x) else 0)
    
    def calc_cycle_point(row):
        r = row['Số ngày mua gần nhất']
        d1 = row['Số ngày mua gần thứ 2']
        d2 = row['Số ngày mua gần thứ 3']
        if pd.isna(r) or pd.isna(d1) or pd.isna(d2):
            return 0
        avg_cycle = (d1 + d2) / 2
        if avg_cycle == 0: return 0
        return min(1, r / avg_cycle)
    
    df_scoring['Điểm Chu kỳ'] = df_scoring.apply(calc_cycle_point, axis=1)
    
    df_scoring['ĐIỂM TỔNG'] = (
        df_scoring['Điểm R'] * w_r +
        df_scoring['Điểm F'] * w_f +
        df_scoring['Điểm M'] * w_m +
        df_scoring['Điểm Hạng'] * w_rank +
        df_scoring['Điểm Call'] * w_call +
        df_scoring['Điểm Chu kỳ'] * w_cycle
    )
    
    # Final cleanup for Cham_diem_KH
    df_chamdiem = df_scoring[['Mã KH', 'Tên khách hàng', 'Hạng KH', 'Điểm R', 'Điểm F', 'Điểm M', 'Điểm Hạng', 'Điểm Call', 'Điểm Chu kỳ', 'ĐIỂM TỔNG', 'Call thiếu', 'Số ngày chưa gặp', 'Ngày gặp cuối']].copy()
    df_chamdiem = df_chamdiem.rename(columns={'Tên khách hàng': 'Tên KH', 'Call thiếu': 'Call còn lại'})
    df_chamdiem['Chọn'] = np.nan
    
    df_chamdiem = df_chamdiem.sort_values(by='ĐIỂM TỔNG', ascending=False)
    
    # Re-order and rename df_tonghop according to user request
    tonghop_cols = [
        'Mã KH', 'Tên khách hàng', 
        'Ngày mua cuối', 'Số lần mua trong 6 tháng', 'Tổng doanh số 6 tháng', 'Số ngày mua gần nhất', 'Số ngày mua gần thứ 2', 'Số ngày mua gần thứ 3', 
        'Số lần đã gặp', 'Số ngày chưa gặp', 
        'Hạng KH', 'Tổng DS tháng', 'Tổng DS tháng trước',
        'Call thiếu', 'DS mục tiêu', 'DS thiếu',
        'Xu hướng mua', 'Chu kỳ TB', 'Dự báo ngày mua tiếp', 'Trạng thái hoạt động'
    ]
    df_tonghop = df_tonghop[tonghop_cols].copy()
    
    # Format dates
    if 'Ngày mua cuối' in df_tonghop.columns:
        df_tonghop['Ngày mua cuối'] = pd.to_datetime(df_tonghop['Ngày mua cuối']).dt.strftime('%d/%m/%Y')
        df_tonghop['Ngày mua cuối'] = df_tonghop['Ngày mua cuối'].replace('NaT', None)
    
    # Save to Excel in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_tonghop.to_excel(writer, sheet_name='7_TongHop', index=False, startrow=1)
        df_chamdiem.to_excel(writer, sheet_name='Cham_diem_KH', index=False, startrow=1)
        df_nhanvien.to_excel(writer, sheet_name='Tong_Hop_Nhan_Vien', index=False)
        
        # Apply visual styling
        style_excel(writer, df_tonghop, df_chamdiem)
    
    output.seek(0)
    return output, df_tonghop, df_chamdiem, df_nhanvien

@app.route('/process', methods=['POST'])
def handle_process():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    try:
        import base64
        # Save temp file
        temp_path = 'temp_input.xlsx'
        file.save(temp_path)
        
        result_excel, df_tonghop, df_chamdiem, df_nhanvien = process_data(temp_path)
        
        os.remove(temp_path)
        
        # Replace NaN/Infinity for JSON compatibility
        df_tonghop = df_tonghop.astype(object).where(pd.notnull(df_tonghop), None)
        df_chamdiem = df_chamdiem.astype(object).where(pd.notnull(df_chamdiem), None)
        df_nhanvien = df_nhanvien.astype(object).where(pd.notnull(df_nhanvien), None)
        
        records_tonghop = df_tonghop.to_dict(orient='records')
        records_chamdiem = df_chamdiem.to_dict(orient='records')
        records_nhanvien = df_nhanvien.to_dict(orient='records')
        
        excel_base64 = base64.b64encode(result_excel.read()).decode('utf-8')
        
        return jsonify({
            'data_tonghop': records_tonghop,
            'data_chamdiem': records_chamdiem,
            'data_nhanvien': records_nhanvien,
            'file_b64': excel_base64,
            'filename': 'Result_Thong_Ke.xlsx'
        })
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print("ERROR IN /process:")
        print(error_details)
        return jsonify({
            'error': str(e),
            'traceback': error_details
        }), 500

@app.route('/config', methods=['GET'])
def get_config():
    try:
        wb = openpyxl.load_workbook(CONFIG_PATH, data_only=True)
        ws = wb['Cấu hình']
        config = {
            'w_r': ws.cell(row=1, column=2).value,
            'w_f': ws.cell(row=2, column=2).value,
            'w_m': ws.cell(row=3, column=2).value,
            'w_rank': ws.cell(row=4, column=2).value,
            'w_call': ws.cell(row=5, column=2).value,
            'w_cycle': ws.cell(row=6, column=2).value,
            'max_recency': ws.cell(row=12, column=2).value,
            'f_window_months': ws.cell(row=13, column=2).value,
            'max_frequency': ws.cell(row=14, column=2).value,
            'max_monetary': ws.cell(row=15, column=2).value,
            'max_call_days': ws.cell(row=16, column=2).value,
            'threshold_vip': ws.cell(row=20, column=2).value,
            'threshold_gold': ws.cell(row=21, column=2).value,
            'threshold_normal': ws.cell(row=22, column=2).value,
            'ds_window_months': ws.cell(row=23, column=2).value,
        }
        return jsonify(config)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/config', methods=['POST'])
def update_config():
    try:
        data = request.json
        wb = openpyxl.load_workbook(CONFIG_PATH)
        ws = wb['Cấu hình']
        
        # Helper to safely format percentages
        def fmt_pct(val):
            if isinstance(val, str) and '%' in val: return val
            try:
                # If they send 0.2, convert to 20%
                float_val = float(val)
                if float_val <= 1.0:
                    return f"{int(float_val * 100)}%"
                return f"{int(float_val)}%"
            except:
                return val

        if 'w_r' in data: ws.cell(row=1, column=2, value=fmt_pct(data['w_r']))
        if 'w_f' in data: ws.cell(row=2, column=2, value=fmt_pct(data['w_f']))
        if 'w_m' in data: ws.cell(row=3, column=2, value=fmt_pct(data['w_m']))
        if 'w_rank' in data: ws.cell(row=4, column=2, value=fmt_pct(data['w_rank']))
        if 'w_call' in data: ws.cell(row=5, column=2, value=fmt_pct(data['w_call']))
        if 'w_cycle' in data: ws.cell(row=6, column=2, value=fmt_pct(data['w_cycle']))
        
        if 'max_recency' in data: ws.cell(row=12, column=2, value=float(data['max_recency']))
        if 'f_window_months' in data: ws.cell(row=13, column=2, value=float(data['f_window_months']))
        if 'max_frequency' in data: ws.cell(row=14, column=2, value=float(data['max_frequency']))
        if 'max_monetary' in data: ws.cell(row=15, column=2, value=float(data['max_monetary']))
        if 'max_call_days' in data: ws.cell(row=16, column=2, value=float(data['max_call_days']))
        
        if 'threshold_vip' in data: ws.cell(row=20, column=2, value=float(data['threshold_vip']))
        if 'threshold_gold' in data: ws.cell(row=21, column=2, value=float(data['threshold_gold']))
        if 'threshold_normal' in data: ws.cell(row=22, column=2, value=float(data['threshold_normal']))
        if 'ds_window_months' in data: ws.cell(row=23, column=2, value=float(data['ds_window_months']))
        
        wb.save(CONFIG_PATH)
        return jsonify({'status': 'success'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/')
def serve_index():
    return app.send_static_file('index.html')

@app.errorhandler(404)
def not_found(e):
    return app.send_static_file('index.html')

if __name__ == '__main__':
    app.run(port=5000, debug=True)
