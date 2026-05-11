from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def style_excel(writer, df_tonghop, df_chamdiem):
    workbook = writer.book
    
    # Define colors
    groups_tonghop = [
        ("THÔNG TIN KH", 2, "EBF5FF", "2563EB"),
        ("LỊCH SỬ MUA HÀNG", 6, "ECFDF5", "059669"),
        ("HOẠT ĐỘNG GHÉ THĂM", 2, "F5F3FF", "7C3AED"),
        ("PHÂN LOẠI & DOANH SỐ THÁNG", 3, "FFFBEB", "D97706"),
        ("CHỈ SỐ THIẾU HỤT", 3, "FEF2F2", "DC2626"),
        ("XU HƯỚNG & HÀNH ĐỘNG", 4, "ECFEFF", "0891B2")
    ]
    
    groups_chamdiem = [
        ("THÔNG TIN KH", 3, "EBF5FF", "2563EB"),
        ("ĐIỂM THÀNH PHẦN RFM", 6, "FFF1F2", "E11D48"),
        ("TỔNG HỢP", 1, "FFFBEB", "D97706"),
        ("HOẠT ĐỘNG", 3, "F5F3FF", "7C3AED"),
        ("CHỌN", 1, "F3F4F6", "4B5563")
    ]
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    def apply_styles(sheet_name, groups, df):
        if sheet_name not in workbook.sheetnames:
            return
        ws = workbook[sheet_name]
        
        # Insert a row at the top for group headers
        ws.insert_rows(1)
        
        col_idx = 1
        for group_name, span, bg_color, text_color in groups:
            start_col = col_idx
            end_col = col_idx + span - 1
            
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col)
            cell.value = group_name
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True, color=text_color)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            
            for c in range(start_col, end_col + 1):
                # Format sub-headers (row 2)
                sub_cell = ws.cell(row=2, column=c)
                sub_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                sub_cell.font = Font(bold=True, color=text_color)
                sub_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                sub_cell.border = thin_border
                ws.cell(row=1, column=c).border = thin_border
            
            col_idx += span
            
        # Freeze panes
        ws.freeze_panes = "A3"
        
        # Adjust column widths
        for col in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
            
        # Format data cells
        for r in range(3, len(df) + 3):
            for c in range(1, len(df.columns) + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                col_name = df.columns[c - 1]
                
                # Borders
                cell.border = thin_border
                
                if val is None:
                    continue
                    
                # Format numbers
                if isinstance(val, (int, float)):
                    if "Điểm" in col_name or col_name == "ĐIỂM TỔNG":
                        cell.number_format = '0.00'
                    elif "DS" in col_name or "doanh số" in col_name.lower():
                        cell.number_format = '#,##0'
                    
                # Rank Colors
                if col_name == 'Hạng KH':
                    if val == 'VIP': cell.fill, cell.font = PatternFill(start_color='FEF08A', end_color='FEF08A', fill_type='solid'), Font(color='854D0E', bold=True)
                    elif val == 'GOLD': cell.fill, cell.font = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid'), Font(color='B45309', bold=True)
                    elif val == 'NORMAL': cell.fill, cell.font = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid'), Font(color='334155', bold=True)
                    elif val == 'NEW': cell.fill, cell.font = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid'), Font(color='1E40AF', bold=True)
                
                # Trạng thái
                elif col_name == 'Trạng thái hoạt động':
                    if val == 'Hoạt động': cell.fill, cell.font = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid'), Font(color='166534')
                    elif val == 'Cảnh báo': cell.fill, cell.font = PatternFill(start_color='FEF08A', end_color='FEF08A', fill_type='solid'), Font(color='854D0E')
                    elif val == 'Ngủ đông': cell.fill, cell.font = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'), Font(color='991B1B')
                    elif val == 'Chưa mua': cell.fill, cell.font = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid'), Font(color='374151')
                
                # Gap / missing colors
                elif col_name in ['Call thiếu', 'DS thiếu', 'Call còn lại']:
                    try:
                        num = float(val)
                        if num > 0: cell.fill, cell.font = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'), Font(color='991B1B')
                        else: cell.fill, cell.font = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'), Font(color='065F46')
                    except: pass
                elif col_name == 'Số ngày chưa gặp':
                    try:
                        num = float(val)
                        if num > 14: cell.fill, cell.font = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'), Font(color='991B1B')
                        else: cell.fill, cell.font = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'), Font(color='065F46')
                    except: pass
                    
                # Gradient for scores (simplified heat map)
                elif ("Điểm" in col_name or col_name == "ĐIỂM TỔNG") and isinstance(val, (int, float)):
                    s = max(0, min(1, float(val)))
                    if s >= 0.75: cell.fill, cell.font = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid'), Font(color='FFFFFF', bold=True)
                    elif s >= 0.5: cell.fill, cell.font = PatternFill(start_color='F97316', end_color='F97316', fill_type='solid'), Font(color='FFFFFF', bold=True)
                    elif s >= 0.25: cell.fill, cell.font = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid'), Font(color='FFFFFF', bold=True)
                    else: cell.fill, cell.font = PatternFill(start_color='64748B', end_color='64748B', fill_type='solid'), Font(color='FFFFFF', bold=True)

    apply_styles('7_TongHop', groups_tonghop, df_tonghop)
    apply_styles('Cham_diem_KH', groups_chamdiem, df_chamdiem)
