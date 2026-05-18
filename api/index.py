from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import io
import math

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from hanh_vi import process_hanh_vi_data

app = Flask(__name__, static_folder='dist', static_url_path='/')
app.config['JSON_SORT_KEYS'] = False
CORS(app)

CONFIG_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'Cau_hinh.xlsx')

def style_excel(writer, df_tonghop, df_chamdiem):
    workbook = writer.book
    groups_tonghop = [
        ("THÔNG TIN KH", 2, "EBF5FF", "2563EB"),
        ("LỊCH SỬ MUA HÀNG", 6, "ECFDF5", "059669"),
        ("HOẠT ĐỘNG GHÉ THĂM", 2, "F5F3FF", "7C3AED"),
        ("PHÂN LOẠI & DOANH SỐ THÁNG", 3, "FFFBEB", "D97706"),
        ("CHỈ SỐ THIẾU HỤT", 3, "FEF2F2", "DC2626"),
        ("XU HƯỚNG & HÀNH ĐỘNG", 4, "ECFEFF", "0891B2")
    ]
    groups_chamdiem = [
        ("THÔNG TIN KH", 3, "EBF5FF", "2563EB"),
        ("ĐIỂM THÀNH PHẦN RFM", 6, "FFF1F2", "E11D48"),
        ("TỔNG HỢP", 1, "FFFBEB", "D97706"),
        ("HOẠT ĐỘNG", 4, "F5F3FF", "7C3AED"),
        ("CHỌN", 1, "F3F4F6", "4B5563")
    ]
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    def apply_styles(sheet_name, groups, df):
        if sheet_name not in workbook.sheetnames: return
        ws = workbook[sheet_name]
        ws.insert_rows(1)
        col_idx = 1
        for group_name, span, bg_color, text_color in groups:
            start_col = col_idx
            end_col = col_idx + span - 1
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col)
            cell.value = group_name
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True, color=text_color)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            for c in range(start_col, end_col + 1):
                sub_cell = ws.cell(row=2, column=c)
                sub_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                sub_cell.font = Font(bold=True, color=text_color)
                sub_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                sub_cell.border = thin_border
                ws.cell(row=1, column=c).border = thin_border
            col_idx += span
        ws.freeze_panes = "C3"
        
        # Auto-fit column widths based on content
        for col in range(1, len(df.columns) + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            
            # Check header length
            header_val = str(ws.cell(row=2, column=col).value or "")
            max_length = len(header_val)
            
            # Check data length (up to 100 rows for performance)
            for row in range(3, min(len(df) + 3, 103)):
                cell_val = str(ws.cell(row=row, column=col).value or "")
                if len(cell_val) > max_length:
                    max_length = len(cell_val)
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = min(max(adjusted_width, 10), 40)

        for r in range(3, len(df) + 3):
            for c in range(1, len(df.columns) + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                col_name = df.columns[c - 1]
                cell.border = thin_border
                if val is None: continue
                if isinstance(val, (int, float)):
                    if "Điểm" in col_name or col_name == "ĐIỂM TỔNG": cell.number_format = '0.00'
                    elif "DS" in col_name or "doanh số" in col_name.lower() or col_name == "Chu kỳ TB": cell.number_format = '#,##0'
                if col_name == 'Hạng KH':
                    if val == 'VIP': cell.fill, cell.font = PatternFill(start_color='FEF08A', end_color='FEF08A', fill_type='solid'), Font(color='854D0E', bold=True)
                    elif val == 'GOLD': cell.fill, cell.font = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid'), Font(color='B45309', bold=True)
                    elif val == 'NORMAL': cell.fill, cell.font = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid'), Font(color='334155', bold=True)
                    elif val == 'NEW': cell.fill, cell.font = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid'), Font(color='1E40AF', bold=True)
                elif col_name == 'Trạng thái hoạt động':
                    if val == 'Hoạt động': cell.fill, cell.font = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid'), Font(color='166534')
                    elif val == 'Cảnh báo': cell.fill, cell.font = PatternFill(start_color='FEF08A', end_color='FEF08A', fill_type='solid'), Font(color='854D0E')
                    elif val == 'Ngủ đông': cell.fill, cell.font = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'), Font(color='991B1B')
                    elif val == 'Chưa mua': cell.fill, cell.font = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid'), Font(color='374151')
                elif col_name in ['Call thiếu', 'DS thiếu', 'Call còn thiếu', 'DS còn thiếu']:
                    try:
                        num = float(val)
                        if num > 0: cell.fill, cell.font = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'), Font(color='991B1B')
                        else: cell.fill, cell.font = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'), Font(color='065F46')
                    except: pass
                elif col_name == 'Số ngày chưa gặp':
                    try:
                        num = float(val)
                        if num > 14: cell.fill, cell.font = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'), Font(color='991B1B')
                        else: cell.fill, cell.font = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'), Font(color='065F46')
                    except: pass
                elif ("Điểm" in col_name or col_name == "ĐIỂM TỔNG") and isinstance(val, (int, float)):
                    s = max(0, min(1, float(val)))
                    if s >= 0.75: cell.fill, cell.font = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid'), Font(color='FFFFFF', bold=True)
                    elif s >= 0.5: cell.fill, cell.font = PatternFill(start_color='F97316', end_color='F97316', fill_type='solid'), Font(color='FFFFFF', bold=True)
                    elif s >= 0.25: cell.fill, cell.font = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid'), Font(color='FFFFFF', bold=True)
                    else: cell.fill, cell.font = PatternFill(start_color='64748B', end_color='64748B', fill_type='solid'), Font(color='FFFFFF', bold=True)
    apply_styles('7_TongHop', groups_tonghop, df_tonghop)
    apply_styles('Cham_diem_KH', groups_chamdiem, df_chamdiem)

def excel_date_to_datetime(excel_date):
    if pd.isna(excel_date):
        return None
    return datetime(1899, 12, 30) + timedelta(days=int(excel_date))

def process_data(input_file_path):
    if not os.path.exists(CONFIG_PATH):
        raise FileNotFoundError(f"Không tìm thấy file cấu hình tại: {CONFIG_PATH}. Vui lòng kiểm tra lại.")
        
    # Load Configuration using openpyxl to avoid pandas header offset issues
    wb = openpyxl.load_workbook(CONFIG_PATH, data_only=True)
    ws = wb['Cấu hình']
    
    def parse_pct(val):
        if isinstance(val, str) and '%' in val:
            return float(val.replace('%', '')) / 100
        return float(val) if val is not None else 0.0

    w_r = parse_pct(ws.cell(row=1, column=2).value)
    w_f = parse_pct(ws.cell(row=2, column=2).value)
    w_m = parse_pct(ws.cell(row=3, column=2).value)
    w_rank = parse_pct(ws.cell(row=4, column=2).value)
    w_call = parse_pct(ws.cell(row=5, column=2).value)
    w_cycle = parse_pct(ws.cell(row=6, column=2).value)
    
    # System Params
    today_excel = ws.cell(row=10, column=2).value
    # Use actual today's date instead of config date to prevent confusion
    today = datetime.now()
    
    max_recency = float(ws.cell(row=12, column=2).value or 120)
    f_window_months = float(ws.cell(row=13, column=2).value or 4)
    max_frequency = float(ws.cell(row=14, column=2).value or 12)
    max_monetary = float(ws.cell(row=15, column=2).value or 18000000)
    max_call_days = float(ws.cell(row=16, column=2).value or 120)
    
    # Rank thresholds
    threshold_vip = float(ws.cell(row=20, column=2).value or 18000000)
    threshold_gold = float(ws.cell(row=21, column=2).value or 6000000)
    threshold_normal = float(ws.cell(row=22, column=2).value or 1000000)
    ds_window_months = float(ws.cell(row=23, column=2).value or 6)
    
    # Load Input Data
    try:
        # Try openpyxl first (for .xlsx)
        xl = pd.ExcelFile(input_file_path, engine='openpyxl')
    except Exception:
        try:
            # Fallback to xlrd (for .xls)
            xl = pd.ExcelFile(input_file_path, engine='xlrd')
        except ImportError:
            raise ImportError("Thiếu thư viện 'xlrd' để đọc file .xls. Vui lòng chạy: pip install xlrd")
        except Exception as e2:
            raise ValueError(f"Lỗi khi đọc file Excel: {str(e2)}")

    with xl:
        required_sheets = ['1_NhanVien', '2_KHTrongTam', '3_DonHang', '4_Call', '5_FrequencyF']
        missing_sheets = [s for s in required_sheets if s not in xl.sheet_names]
        if missing_sheets:
            raise ValueError(f"File Excel thiếu các sheet bắt buộc: {', '.join(missing_sheets)}")

        df_nhanvien = pd.read_excel(xl, '1_NhanVien', dtype=str)
        df_kh = pd.read_excel(xl, '2_KHTrongTam')
        df_orders = pd.read_excel(xl, '3_DonHang')
        df_calls = pd.read_excel(xl, '4_Call')
        df_freq = pd.read_excel(xl, '5_FrequencyF')
    
    # Enforce string type for IDs to ensure correct merging
    for df in [df_nhanvien, df_kh, df_orders, df_calls, df_freq]:
        for col in ['Mã KH', 'Mã KHTC']:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()
    
    # Filter out customers with ID length > 9 (as requested)
    df_kh = df_kh[df_kh['Mã KH'].str.len() <= 9]
    
    # Ensure date formats - Optimized for yyyy-MM-dd HH:mm:ss (standard ISO)
    df_orders['Ngày đặt'] = pd.to_datetime(df_orders['Ngày đặt'], errors='coerce')
    df_calls['Thời gian checkin'] = pd.to_datetime(df_calls['Thời gian checkin'], errors='coerce')
    
    # Calculate 6-month window
    start_date_6m = today - timedelta(days=30*ds_window_months)
    
    # --- Processing 7_TongHop ---
    
    # Order statistics
    if 'Số lượng' not in df_orders.columns or 'Đơn giá' not in df_orders.columns:
        df_orders['Revenue'] = 0
    else:
        df_orders['Số lượng'] = pd.to_numeric(df_orders['Số lượng'], errors='coerce').fillna(0)
        df_orders['Đơn giá'] = pd.to_numeric(df_orders['Đơn giá'], errors='coerce').fillna(0)
        df_orders['Revenue'] = df_orders['Số lượng'] * df_orders['Đơn giá']
        
    df_orders_6m = df_orders[df_orders['Ngày đặt'] >= start_date_6m]
    
    # Determine "Current Month" for statistics based on data
    # Determine "Current Month" for statistics based on the latest data in the file
    if not df_orders.empty and pd.notna(df_orders['Ngày đặt'].max()):
        latest_date = df_orders['Ngày đặt'].max()
        current_month = latest_date.month
        current_year = latest_date.year
    else:
        current_month = today.month
        current_year = today.year
        
    prev_month = 12 if current_month == 1 else current_month - 1
    prev_year = current_year - 1 if current_month == 1 else current_year

    order_stats = []
    for makh, group in df_orders_6m.groupby('Mã KH'):
        sorted_dates = sorted(group['Ngày đặt'].dt.date.unique(), reverse=True)
        last_buy = sorted_dates[0] if sorted_dates else None
        
        diff1 = (sorted_dates[0] - sorted_dates[1]).days if len(sorted_dates) > 1 else None
        diff2 = (sorted_dates[1] - sorted_dates[2]).days if len(sorted_dates) > 2 else None
        
        ds_thang = group[(group['Ngày đặt'].dt.month == current_month) & (group['Ngày đặt'].dt.year == current_year)]['Revenue'].sum()
        ds_thang_truoc = group[(group['Ngày đặt'].dt.month == prev_month) & (group['Ngày đặt'].dt.year == prev_year)]['Revenue'].sum()
        
        if ds_thang > ds_thang_truoc: xu_huong = "📈 Tăng"
        elif ds_thang < ds_thang_truoc: xu_huong = "📉 Giảm"
        else: xu_huong = "➡ Ổn định"
        
        days_since = (today.date() - last_buy).days if last_buy else None
        
        chu_ky_tb = None
        if diff1 is not None and diff2 is not None: 
            chu_ky_tb = math.ceil((diff1 + diff2) / 2)
        
        du_bao = None
        if chu_ky_tb is not None:
            du_bao = (today.date() + timedelta(days=chu_ky_tb)).strftime('%d/%m/%Y')
            
        if days_since is None: trang_thai = "Chưa mua"
        elif days_since <= 60: trang_thai = "Hoạt động"
        elif days_since <= 120: trang_thai = "Cảnh báo"
        else: trang_thai = "Ngủ đông"
        
        order_stats.append({
            'Mã KH': makh,
            'Ngày mua cuối': last_buy,
            'Số lần mua trong 6 tháng': len(sorted_dates),
            'Tổng doanh số 6 tháng': group['Revenue'].sum(),
            'Tổng DS tháng': ds_thang,
            'Tổng DS tháng trước': ds_thang_truoc,
            'Xu hướng mua': xu_huong,
            'Số ngày mua gần nhất': days_since,
            'Số ngày mua gần thứ 2': diff1,
            'Số ngày mua gần thứ 3': diff2,
            'Chu kỳ TB': chu_ky_tb,
            'Dự báo ngày mua tiếp': du_bao,
            'Trạng thái hoạt động': trang_thai
        })
    df_order_summary = pd.DataFrame(order_stats)
    
    # Call statistics
    call_stats = []
    for makh, group in df_calls.groupby('Mã KH'):
        last_call = group['Thời gian checkin'].max()
        # Handle cases where all calls have no timestamp (NaT)
        days_since_call = (today - last_call).days if pd.notna(last_call) else None
        
        call_stats.append({
            'Mã KH': makh,
            'Số lần đã gặp': len(group),
            'Số ngày chưa gặp': days_since_call,
            'Ngày gặp cuối': last_call.strftime('%d/%m/%Y') if pd.notna(last_call) else None
        })
    df_call_summary = pd.DataFrame(call_stats)
    
    # Merge everything into 7_TongHop
    df_tonghop = df_kh[['Mã KH', 'Tên KH']].copy()
    df_tonghop = df_tonghop.merge(df_order_summary, on='Mã KH', how='left')
    df_tonghop = df_tonghop.merge(df_call_summary, on='Mã KH', how='left')
    
    # Ranking
    def get_rank(ds):
        if ds >= threshold_vip: return 'VIP'
        if ds >= threshold_gold: return 'GOLD'
        if ds >= threshold_normal: return 'NORMAL'
        return 'NEW'
    
    df_tonghop['Hạng KH'] = df_tonghop['Tổng doanh số 6 tháng'].fillna(0).apply(get_rank)
    
    # FrequencyF merge (Group by Mã KHTC)
    if 'F' in df_freq.columns:
        if 'Call' in df_freq.columns:
            df_freq['Call'] = df_freq[['Call', 'F']].max(axis=1)
        else:
            df_freq['Call'] = df_freq['F']
        
    if 'Mã KHTC' in df_freq.columns:
        df_freq_grouped = df_freq.groupby('Mã KHTC').agg({
            'Doanh số KH': 'max',
            'Call': 'sum'
        }).reset_index().rename(columns={'Mã KHTC': 'Mã KH', 'Doanh số KH': 'DS mục tiêu', 'Call': 'Call mục tiêu'})
    else:
        df_freq_grouped = pd.DataFrame(columns=['Mã KH', 'DS mục tiêu', 'Call mục tiêu'])
        
    df_tonghop = df_tonghop.merge(df_freq_grouped, on='Mã KH', how='left')
    
    # Fill Nans for missing columns after merge
    df_tonghop['Tổng DS tháng'] = df_tonghop['Tổng DS tháng'].fillna(0)
    df_tonghop['Tổng DS tháng trước'] = df_tonghop['Tổng DS tháng trước'].fillna(0)
    df_tonghop['Xu hướng mua'] = df_tonghop['Xu hướng mua'].fillna("Chưa mua")
    df_tonghop['Trạng thái hoạt động'] = df_tonghop['Trạng thái hoạt động'].fillna("Chưa mua")
    df_tonghop['Số lần mua trong 6 tháng'] = df_tonghop['Số lần mua trong 6 tháng'].fillna(0)
    df_tonghop['Tổng doanh số 6 tháng'] = df_tonghop['Tổng doanh số 6 tháng'].fillna(0)
    
    df_tonghop['DS thiếu'] = df_tonghop['DS mục tiêu'].fillna(0) - df_tonghop['Tổng DS tháng']
    df_tonghop['Call thiếu'] = (df_tonghop['Call mục tiêu'].fillna(0) - df_tonghop['Số lần đã gặp'].fillna(0)).apply(lambda x: max(0, x))
    
    # Clean column names to match sample
    df_tonghop = df_tonghop.rename(columns={'Tên KH': 'Tên khách hàng'})
    
    # --- Processing Cham_diem_KH ---
    df_scoring = df_tonghop.copy()
    
    # Scoring Logic
    df_scoring['Điểm R'] = df_scoring['Số ngày mua gần nhất'].apply(lambda x: max(0, 1 - x/max_recency) if pd.notna(x) else 0)
    df_scoring['Điểm F'] = df_scoring['Số lần mua trong 6 tháng'].apply(lambda x: min(1, x/max_frequency) if pd.notna(x) else 0)
    df_scoring['Điểm M'] = df_scoring['Tổng doanh số 6 tháng'].apply(lambda x: min(1, x/max_monetary) if pd.notna(x) else 0)
    
    rank_points = {'VIP': 1.0, 'GOLD': 0.8, 'NORMAL': 0.6, 'NEW': 0.4}
    df_scoring['Điểm Hạng'] = df_scoring['Hạng KH'].map(rank_points)
    
    df_scoring['Điểm Call'] = df_scoring['Số ngày chưa gặp'].apply(lambda x: max(0, 1 - x/max_call_days) if pd.notna(x) else 0)
    
    def calc_cycle_point(row):
        r = row['Số ngày mua gần nhất']
        d1 = row['Số ngày mua gần thứ 2']
        d2 = row['Số ngày mua gần thứ 3']
        if pd.isna(r) or pd.isna(d1) or pd.isna(d2):
            return 0
        avg_cycle = math.ceil((d1 + d2) / 2)
        if avg_cycle == 0: return 0
        return min(1, r / avg_cycle)
    
    df_scoring['Điểm Chu kỳ'] = df_scoring.apply(calc_cycle_point, axis=1)
    
    df_scoring['ĐIỂM TỔNG'] = (
        df_scoring['Điểm R'] * w_r +
        df_scoring['Điểm F'] * w_f +
        df_scoring['Điểm M'] * w_m +
        df_scoring['Điểm Hạng'] * w_rank +
        df_scoring['Điểm Call'] * w_call +
        df_scoring['Điểm Chu kỳ'] * w_cycle
    )
    
    # Final cleanup for Cham_diem_KH
    df_chamdiem = df_scoring[['Mã KH', 'Tên khách hàng', 'Hạng KH', 'Điểm R', 'Điểm F', 'Điểm M', 'Điểm Hạng', 'Điểm Call', 'Điểm Chu kỳ', 'ĐIỂM TỔNG', 'DS thiếu', 'Call thiếu', 'Số ngày chưa gặp', 'Ngày gặp cuối']].copy()
    df_chamdiem = df_chamdiem.rename(columns={'Tên khách hàng': 'Tên KH', 'DS thiếu': 'DS còn thiếu', 'Số ngày chưa gặp': 'Ngày chưa gặp'})
    df_chamdiem['Chọn'] = np.nan
    
    df_chamdiem = df_chamdiem.sort_values(by='ĐIỂM TỔNG', ascending=False)
    
    # Re-order and rename df_tonghop according to user request
    df_tonghop = df_tonghop.rename(columns={
        'Số lần mua trong 6 tháng': 'Số lần mua 6T',
        'Tổng doanh số 6 tháng': 'Tổng DS 6T',
        'Số ngày mua gần nhất': 'Ngày mua gần nhất',
        'Số ngày mua gần thứ 2': 'Ngày mua gần 2',
        'Số ngày mua gần thứ 3': 'Ngày mua gần 3',
        'Tổng DS tháng': 'DS tháng',
        'Tổng DS tháng trước': 'DS tháng trước',
        'Xu hướng mua': 'Xu hướng',
        'Dự báo ngày mua tiếp': 'Dự báo mua',
        'Trạng thái hoạt động': 'Trạng thái',
        'Số ngày chưa gặp': 'Ngày chưa gặp'
    })
    
    tonghop_cols = [
        'Mã KH', 'Tên khách hàng', 
        'Ngày mua cuối', 'Số lần mua 6T', 'Tổng DS 6T', 'Ngày mua gần nhất', 'Ngày mua gần 2', 'Ngày mua gần 3', 
        'Số lần đã gặp', 'Ngày chưa gặp', 
        'Hạng KH', 'DS tháng', 'DS tháng trước',
        'Call thiếu', 'DS mục tiêu', 'DS thiếu',
        'Xu hướng', 'Chu kỳ TB', 'Dự báo mua', 'Trạng thái'
    ]
    df_tonghop = df_tonghop[tonghop_cols].copy()
    
    # Format dates
    if 'Ngày mua cuối' in df_tonghop.columns:
        df_tonghop['Ngày mua cuối'] = pd.to_datetime(df_tonghop['Ngày mua cuối']).dt.strftime('%d/%m/%Y')
        df_tonghop['Ngày mua cuối'] = df_tonghop['Ngày mua cuối'].replace('NaT', None)
    
    # Save to Excel in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_tonghop.to_excel(writer, sheet_name='7_TongHop', index=False, startrow=1)
        df_chamdiem.to_excel(writer, sheet_name='Cham_diem_KH', index=False, startrow=1)
        df_nhanvien.to_excel(writer, sheet_name='Tong_Hop_Nhan_Vien', index=False)
        
        # Apply visual styling
        style_excel(writer, df_tonghop, df_chamdiem)
    
    # --- Calculate Monthly Statistics ---
    month_orders = df_orders[(df_orders['Ngày đặt'].dt.month == current_month) & (df_orders['Ngày đặt'].dt.year == current_year)]
    month_calls = df_calls[(df_calls['Thời gian checkin'].dt.month == current_month) & (df_calls['Thời gian checkin'].dt.year == current_year)]
    
    # Define target customers set for filtering
    target_ids = set(df_kh['Mã KH'].astype(str).unique())
    
    # Filter monthly data to only include target customers
    month_orders_target = month_orders[month_orders['Mã KH'].astype(str).isin(target_ids)]
    month_calls_target = month_calls[month_calls['Mã KH'].astype(str).isin(target_ids)]

    customers_with_orders = set(month_orders_target['Mã KH'].unique())
    customers_with_calls = set(month_calls_target['Mã KH'].unique())
    
    def safe_int(val):
        try:
            return int(val) if pd.notna(val) else 0
        except:
            return 0

    def safe_float(val):
        try:
            f = float(val)
            return f if math.isfinite(f) else 0.0
        except:
            return 0.0

    stats = {
        'total_customers': safe_int(len(df_kh)),
        'customers_with_orders': safe_int(len(customers_with_orders)),
        'customers_with_calls': safe_int(len(customers_with_calls)),
        'customers_with_both': safe_int(len(customers_with_orders.intersection(customers_with_calls))),
        'customers_order_no_call': safe_int(len(customers_with_orders - customers_with_calls)),
        'total_revenue': safe_float(month_orders_target['Revenue'].sum()),
        'total_orders': safe_int(len(month_orders_target.groupby(['Mã KH', 'Ngày đặt'])) if not month_orders_target.empty else 0),
        'total_calls': safe_int(len(month_calls_target.groupby(['Mã KH', 'Thời gian checkin'])) if not month_calls_target.empty else 0),
        'total_calls_all_time': safe_int(len(df_calls[df_calls['Mã KH'].astype(str).isin(target_ids)])),
        'month': current_month,
        'year': current_year
    }

    # --- Identify Casual Customers (Khách vãng lai) ---
    target_makh_set = set(df_kh['Mã KH'].astype(str).unique())
    
    # Casual Orders
    month_orders_casual = month_orders[~month_orders['Mã KH'].astype(str).isin(target_makh_set)]
    casual_order_stats = []
    for makh, group in month_orders_casual.groupby('Mã KH'):
        # Try to find a name in the group
        ten_kh = "Không rõ"
        for col in ['Tên KH', 'Tên khách hàng', 'Tên']:
            if col in group.columns and not pd.isna(group[col].iloc[0]):
                ten_kh = group[col].iloc[0]
                break
                
        last_date = group['Ngày đặt'].max()
        casual_order_stats.append({
            'Mã KH': makh,
            'Tên KH': ten_kh,
            'Tổng DS tháng': group['Revenue'].sum(),
            'Số đơn hàng': len(group),
            'Ngày mua cuối': last_date.strftime('%d/%m/%Y') if pd.notna(last_date) else None
        })
    df_casual_orders = pd.DataFrame(casual_order_stats)
    
    # Casual Calls
    month_calls_casual = month_calls[~month_calls['Mã KH'].astype(str).isin(target_makh_set)]
    casual_call_stats = []
    for makh, group in month_calls_casual.groupby('Mã KH'):
        # Try to find a name in the group
        ten_kh = "Không rõ"
        for col in ['Tên KH', 'Tên khách hàng', 'Tên']:
            if col in group.columns and not pd.isna(group[col].iloc[0]):
                ten_kh = group[col].iloc[0]
                break

        casual_call_stats.append({
            'Mã KH': makh,
            'Tên KH': ten_kh,
            'Số lần Call': len(group),
            'Ngày Call cuối': group['Thời gian checkin'].max().strftime('%d/%m/%Y')
        })
    df_casual_calls = pd.DataFrame(casual_call_stats)
    
    # Add casual stats to the stats object
    stats['casual_with_orders'] = safe_int(len(df_casual_orders))
    stats['casual_with_calls'] = safe_int(len(df_casual_calls))

    output.seek(0)
    return output, df_tonghop, df_chamdiem, df_nhanvien, stats, df_casual_orders, df_casual_calls

@app.route('/process', methods=['POST'])
def handle_process():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    try:
        import base64
        import tempfile
        # Save temp file to /tmp directory which is writable on Vercel
        temp_path = os.path.join(tempfile.gettempdir(), 'temp_input.xlsx')
        file.save(temp_path)
        
        result_excel, df_tonghop, df_chamdiem, df_nhanvien, stats, df_casual_orders, df_casual_calls = process_data(temp_path)
        
        os.remove(temp_path)
        
        # Replace NaN/Infinity for JSON compatibility
        df_tonghop = df_tonghop.astype(object).where(pd.notnull(df_tonghop), None)
        df_chamdiem = df_chamdiem.astype(object).where(pd.notnull(df_chamdiem), None)
        df_nhanvien = df_nhanvien.astype(object).where(pd.notnull(df_nhanvien), None)
        df_casual_orders = df_casual_orders.astype(object).where(pd.notnull(df_casual_orders), None)
        df_casual_calls = df_casual_calls.astype(object).where(pd.notnull(df_casual_calls), None)
        
        records_tonghop = df_tonghop.to_dict(orient='records')
        records_chamdiem = df_chamdiem.to_dict(orient='records')
        records_nhanvien = df_nhanvien.to_dict(orient='records')
        
        excel_base64 = base64.b64encode(result_excel.read()).decode('utf-8')
        
        return jsonify({
            'data_tonghop': records_tonghop,
            'data_chamdiem': records_chamdiem,
            'data_nhanvien': records_nhanvien,
            'data_casual_orders': df_casual_orders.to_dict(orient='records') if not df_casual_orders.empty else [],
            'data_casual_calls': df_casual_calls.to_dict(orient='records') if not df_casual_calls.empty else [],
            'stats': stats,
            'file_b64': excel_base64,
            'filename': 'Result_Thong_Ke.xlsx'
        })
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print("ERROR IN /process:")
        # Use repr() or encode/decode to avoid UnicodeEncodeError in Windows console
        try:
            print(error_details)
        except UnicodeEncodeError:
            print(error_details.encode('ascii', 'replace').decode())
            
        return jsonify({
            'error': str(e),
            'traceback': error_details
        }), 500

@app.route('/config', methods=['GET'])
def get_config():
    try:
        wb = openpyxl.load_workbook(CONFIG_PATH, data_only=True)
        ws = wb['Cấu hình']
        config = {
            'w_r': ws.cell(row=1, column=2).value,
            'w_f': ws.cell(row=2, column=2).value,
            'w_m': ws.cell(row=3, column=2).value,
            'w_rank': ws.cell(row=4, column=2).value,
            'w_call': ws.cell(row=5, column=2).value,
            'w_cycle': ws.cell(row=6, column=2).value,
            'max_recency': ws.cell(row=12, column=2).value,
            'f_window_months': ws.cell(row=13, column=2).value,
            'max_frequency': ws.cell(row=14, column=2).value,
            'max_monetary': ws.cell(row=15, column=2).value,
            'max_call_days': ws.cell(row=16, column=2).value,
            'threshold_vip': ws.cell(row=20, column=2).value,
            'threshold_gold': ws.cell(row=21, column=2).value,
            'threshold_normal': ws.cell(row=22, column=2).value,
            'ds_window_months': ws.cell(row=23, column=2).value,
        }
        return jsonify(config)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/config', methods=['POST'])
def update_config():
    try:
        data = request.json
        wb = openpyxl.load_workbook(CONFIG_PATH)
        ws = wb['Cấu hình']
        
        # Helper to safely format percentages
        def fmt_pct(val):
            if isinstance(val, str) and '%' in val: return val
            try:
                # If they send 0.2, convert to 20%
                float_val = float(val)
                if float_val <= 1.0:
                    return f"{int(float_val * 100)}%"
                return f"{int(float_val)}%"
            except:
                return val

        if 'w_r' in data: ws.cell(row=1, column=2, value=fmt_pct(data['w_r']))
        if 'w_f' in data: ws.cell(row=2, column=2, value=fmt_pct(data['w_f']))
        if 'w_m' in data: ws.cell(row=3, column=2, value=fmt_pct(data['w_m']))
        if 'w_rank' in data: ws.cell(row=4, column=2, value=fmt_pct(data['w_rank']))
        if 'w_call' in data: ws.cell(row=5, column=2, value=fmt_pct(data['w_call']))
        if 'w_cycle' in data: ws.cell(row=6, column=2, value=fmt_pct(data['w_cycle']))
        
        if 'max_recency' in data: ws.cell(row=12, column=2, value=float(data['max_recency']))
        if 'f_window_months' in data: ws.cell(row=13, column=2, value=float(data['f_window_months']))
        if 'max_frequency' in data: ws.cell(row=14, column=2, value=float(data['max_frequency']))
        if 'max_monetary' in data: ws.cell(row=15, column=2, value=float(data['max_monetary']))
        if 'max_call_days' in data: ws.cell(row=16, column=2, value=float(data['max_call_days']))
        
        if 'threshold_vip' in data: ws.cell(row=20, column=2, value=float(data['threshold_vip']))
        if 'threshold_gold' in data: ws.cell(row=21, column=2, value=float(data['threshold_gold']))
        if 'threshold_normal' in data: ws.cell(row=22, column=2, value=float(data['threshold_normal']))
        if 'ds_window_months' in data: ws.cell(row=23, column=2, value=float(data['ds_window_months']))
        
        wb.save(CONFIG_PATH)
        return jsonify({'status': 'success'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/process_hanh_vi', methods=['POST'])
def handle_process_hanh_vi():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    try:
        import tempfile
        import os
        import uuid
        import gzip

        temp_path = os.path.join(tempfile.gettempdir(), f'temp_hanhvi_{uuid.uuid4().hex}.xlsx')

        is_compressed = request.headers.get('x-content-encoding') == 'gzip'

        if is_compressed:
            # Decompress gzip-encoded file sent from frontend
            compressed_data = file.read()
            decompressed_data = gzip.decompress(compressed_data)
            with open(temp_path, 'wb') as f:
                f.write(decompressed_data)
        else:
            file.save(temp_path)
        
        result = process_hanh_vi_data(temp_path)
        
        try:
            os.remove(temp_path)
        except Exception as e:
            print("Could not remove temp file:", e)
        
        if result.get("Status") == "Error":
            return jsonify({'error': result.get("Message"), 'traceback': result.get("Traceback")}), 500
            
        return jsonify(result)
        
    except Exception as e:
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@app.route('/ping')
def ping():
    return 'pong'

@app.route('/')
def serve_index():
    return app.send_static_file('index.html')

@app.errorhandler(404)
def not_found(e):
    return app.send_static_file('index.html')

if __name__ == '__main__':
    app.run(port=5000, debug=True)
